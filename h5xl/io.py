"""Convert HDF5 tables into Excel worksheets."""

import openpyxl as xl
from contextlib import contextmanager
import numpy as np


def record_iter(record):
    """Iterate over numpy records, flatten composite types into simple types"""
    for item in record:
        if isinstance(item, np.void):
            for result in record_iter(item):
                yield result
        else:
            yield item


def flatten_dtype_names(dtype, name=""):
    """Flattens compound dtype names into dot separated names"""
    if dtype.fields:
        for subname in dtype.names:
            (subdtype, _) = dtype.fields[subname]

            if name:
                subname = ".".join((name, subname))

            for n in flatten_dtype_names(subdtype, subname):
                yield n

    else:
        yield name


class ExcelWriter(object):

    def __init__(self, wb):
        self.wb = wb
        self.ws = {}

    def write(self, ndarray, sheet_name):
        if len(ndarray.shape) == 1 and not ndarray.dtype.fields:
            self._write_1d(ndarray, sheet_name)

        elif (len(ndarray.shape) == 2
              or (len(ndarray.shape) == 1 and ndarray.dtype.fields)):
            self._write_2d(ndarray, sheet_name)

        else:
            raise ValueError("Dataset must be at most two dimensions.")

    def writerow(self, row, sheet_name):
        if sheet_name not in self.ws:
            self.ws[sheet_name] = self.wb.create_sheet(title=sheet_name)

        self.ws[sheet_name].append(row)

    def _write_1d(self, ndarray, sheet_name):
        for val in ndarray:
            self.writerow([val], sheet_name)


    def _write_2d(self, ndarray, sheet_name):
        if ndarray.dtype.fields:
            self.writerow(list(flatten_dtype_names(ndarray.dtype)), sheet_name)

        for row in ndarray:
            self.writerow(list(record_iter(row)), sheet_name)




@contextmanager
def writer(path, mode='w'):
    if mode == 'w':
        wb = xl.Workbook()
    elif mode == 'a':
        wb = xl.load_workbook(path)
    else:
        raise ValueError("'{}' is not a valid mode.".format(mode))

    yield ExcelWriter(wb)
    wb.save(path)
