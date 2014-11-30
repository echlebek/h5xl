from unittest import TestCase
import numpy as np
from h5xl.io import ExcelWriter
import tempfile
import openpyxl as xl


def fmt_sheet(ws):
    return [[c.value for c in row] for row in ws.rows]


class TestIO(TestCase):

    def test_write_1d(self):
        a = np.array([1,2,3,4])

        wb = xl.Workbook()
        w = ExcelWriter(wb)
        w.write(a, "test")
        self.assertEquals(fmt_sheet(wb["test"]), [[1], [2], [3], [4]])

    def test_flatten_dtype_names(self):
        from h5xl.io import flatten_dtype_names

        d = np.dtype([("a", np.dtype([("B", int), ("C", int)])), ("b", int)])
        names = flatten_dtype_names(d)

        self.assertEquals(list(names), ["a.B", "a.C", "b"])

    def test_write_2d(self):
        a = np.array([[1, 2, 3, 4], [2, 3, 4, 5]])

        wb = xl.Workbook()
        w = ExcelWriter(wb)
        w.write(a, "test")
        self.assertEquals(fmt_sheet(wb["test"]), [[1, 2, 3, 4], [2, 3, 4, 5]])

    def test_write_2d_record_array(self):
        a = np.array([(1, 2, 3, 4), (2, 3, 4, 5)], dtype=np.dtype(
            [("a", int), ("b", float), ("c", int), ("d", float)]
        ))

        wb = xl.Workbook()
        w = ExcelWriter(wb)
        w.write(a, "test")
        self.assertEquals(
            fmt_sheet(wb["test"]),
            [[u"a", u"b", u"c", u"d"], [1, 2.0, 3, 4.0], [2, 3.0, 4, 5.0]]
        )


if __name__ == "__main__":
    unittest.main()
